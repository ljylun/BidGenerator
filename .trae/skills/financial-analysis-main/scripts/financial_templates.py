#!/usr/bin/env python3
"""
Financial Analysis Helper Scripts
Generates Excel templates for financial modeling
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_comps_template():
    """Create a comparable company analysis Excel template"""
    wb = openpyxl.Workbook()
    
    # Operating Metrics sheet
    ws1 = wb.active
    ws1.title = "Operating Metrics"
    
    # Headers
    headers = ["Company", "Revenue", "Revenue Growth", "Gross Profit", "Gross Margin", 
               "EBITDA", "EBITDA Margin", "Net Income", "Net Margin"]
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="17365D", end_color="17365D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Sample companies (3 rows)
    for row in range(2, 5):
        ws1.cell(row=row, column=1, value=f"Company {row-1}")
    
    # Statistics rows
    stats_row = 6
    for col in range(2, len(headers) + 1):
        col_letter = get_column_letter(col)
        ws1.cell(row=stats_row, column=col, value=f"=MEDIAN({col_letter}2:{col_letter}4)")
    
    # Valuation sheet
    ws2 = wb.create_sheet("Valuation")
    val_headers = ["Company", "Market Cap", "Enterprise Value", "EV/Revenue", "EV/EBITDA", "P/E"]
    
    for col, header in enumerate(val_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="17365D", end_color="17365D", fill_type="solid")
    
    # Save
    wb.save("comps_template.xlsx")
    print("Created comps_template.xlsx")

def create_dcf_template():
    """Create a DCF model Excel template"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DCF Model"
    
    # Headers
    headers = ["Year", "Revenue", "Growth", "EBITDA", "EBITDA Margin", "D&A", "EBIT", 
               "Tax", "NOPAT", "CapEx", "Δ WC", "FCF", "Discount Factor", "PV of FCF"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
    
    # 5 year projection
    for row in range(2, 7):
        ws.cell(row=row, column=1, value=f"Year {row-1}")
    
    # Terminal value row
    ws.cell(row=8, column=1, value="Terminal Value")
    ws.cell(row=9, column=1, value="Enterprise Value")
    ws.cell(row=10, column=1, value="Less: Net Debt")
    ws.cell(row=11, column=1, value="Equity Value")
    
    wb.save("dcf_template.xlsx")
    print("Created dcf_template.xlsx")

if __name__ == "__main__":
    create_comps_template()
    create_dcf_template()
    print("Templates created successfully!")
