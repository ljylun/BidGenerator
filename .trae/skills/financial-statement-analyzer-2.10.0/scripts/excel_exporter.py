#!/usr/bin/env python3
"""Excel导出器 — 将分析结果导出为结构化Excel摘要。

功能: 生成含KPI/比率/风险信号/建议的Excel汇总表。
"""
from __future__ import annotations
import logging
import os
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


def export_to_excel(
    results: Dict[str, Any],
    output_path: str,
    company_name: str = "",
) -> str:
    """导出分析结果为Excel摘要。

    Args:
        results: 分析结果字典
        output_path: 输出Excel文件路径
        company_name: 公司/集团名称

    Returns:
        输出文件路径
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed")
        return ""

    wb = openpyxl.Workbook()

    # 样式
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="26215C", end_color="26215C", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    num_align = Alignment(horizontal="right")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    red_font = Font(color="A32D2D")
    green_font = Font(color="3B6D11")

    # ===== Sheet 1: KPI & Ratios =====
    ws1 = wb.active
    ws1.title = "KPI与比率"
    _write_header_row(ws1, 1, ["指标", "数值", "评级", "说明"], header_font, header_fill, header_align, thin_border)

    ratios = results.get("ratios", {})
    row = 2
    # 盈利能力
    for cat_name, cat_data in [("盈利能力", "盈利能力"), ("偿债能力", "偿债能力"), ("营运能力", "营运能力"), ("成长能力", "成长能力")]:
        cat = ratios.get(cat_name, [])
        if cat:
            _write_cell(ws1, row, 1, f"【{cat_name}】", Font(bold=True, size=11))
            row += 1
            latest = cat[-1]
            for k, v in latest.items():
                if k == "period":
                    continue
                display = f"{v*100:.1f}%" if isinstance(v, float) and abs(v) < 10 else f"{v:.2f}"
                _write_row(ws1, row, [k, display, "", ""], thin_border, num_align)
                row += 1

    # ===== Sheet 2: Red Flags =====
    ws2 = wb.create_sheet("风险信号")
    _write_header_row(ws2, 1, ["#", "级别", "信号名称", "风险类型", "描述"], header_font, header_fill, header_align, thin_border)

    redflags = results.get("redflags", {})
    signals = redflags.get("triggered_signals", [])
    row = 2
    for i, s in enumerate(signals, 1):
        _write_row(ws2, row, [i, s["severity"], s["name"], s.get("risk_type", ""), s.get("description", "")], thin_border)
        if s["severity"] == "P0":
            ws2.cell(row=row, column=2).font = red_font
        row += 1

    # ===== Sheet 3: Advice =====
    ws3 = wb.create_sheet("建议清单")
    _write_header_row(ws3, 1, ["#", "优先级", "场景", "行动", "负责人", "截止时间"], header_font, header_fill, header_align, thin_border)

    advice = results.get("advice", {})
    checklist = advice.get("action_checklist", [])
    row = 2
    for i, item in enumerate(checklist, 1):
        _write_row(ws3, row, [i, item.get("priority", ""), item.get("scenario", ""), item.get("task", ""), item.get("owner", ""), item.get("deadline", "")], thin_border)
        row += 1

    # ===== Sheet 4: Data Quality =====
    ws4 = wb.create_sheet("数据质量")
    _write_header_row(ws4, 1, ["项目", "详情"], header_font, header_fill, header_align, thin_border)

    qs = results.get("data_quality", {})
    row = 2
    _write_row(ws4, row, ["综合评分", f"{qs.get('total_score', 'N/A')}/100"], thin_border); row += 1
    _write_row(ws4, row, ["置信度", qs.get("confidence", "N/A")], thin_border); row += 1
    for dim, info in qs.get("breakdown", {}).items():
        detail = info.get("detail", "")
        if isinstance(detail, list):
            detail = "; ".join(detail)
        _write_row(ws4, row, [dim, f"{info.get('score','?')}/{info.get('max','?')} - {detail}"], thin_border)
        row += 1

    # 列宽自适应
    for ws in [ws1, ws2, ws3, ws4]:
        for col_idx in range(1, 7):
            max_len = 0
            for cell in ws[get_column_letter(col_idx)]:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    wb.save(output_path)
    logger.info("Excel exported to %s", output_path)
    return output_path


def _write_header_row(ws, row, headers, font, fill, align, border):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
        cell.border = border


def _write_row(ws, row, values, border, num_align=None):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.border = border
        if isinstance(v, (int, float)) and num_align:
            cell.alignment = num_align


def _write_cell(ws, row, col, value, font=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
